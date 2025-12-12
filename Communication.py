import os
import urllib
import threading

import socket
import logging
import time as ti
import requests
import hashlib
import uuid
import time
import binascii

import Constant as Cons
import MainFunction as Mf
import Response as Res
import System_Info as SysInfo
import DRS_Response as DR_r
import TTL_Communication as ttl
import Calculate_CMD as Cal
import ASYNC_Temp as Async

from socket import AF_INET, SOCK_STREAM
from requests.auth import HTTPBasicAuth
from requests.auth import HTTPDigestAuth
from datetime import datetime
from openpyxl import Workbook, load_workbook
from icecream import ic

import Dialog

file_lock = threading.Lock()


# 2025.06.26 Added Function for TMS-10
def send_cmd_for_multi(send_cmd, title, root_view):
    find_ch()
    host = Cons.selected_ch['ip']
    input_port = Cons.selected_ch['port']
    port = int(0) if Cons.port == '' else int(input_port)
    buf_size = Cons.buf_size
    client = socket.socket(AF_INET, SOCK_STREAM)

    try:
        client.settimeout(3)
        # -*- coding: utf-8 -*-
        client.connect((host, port))
        if Cons.data_sending:
            client.send(bytes(send_cmd))
            reply = client.recv(buf_size)
            hex_data = binascii.hexlify(reply).decode('utf-8')
            ic(hex_data)
        else:
            print('Protocol sending was stopped')

    except socket.error as err:
        print(f'network error:{err}')
        dialog_txt = f'Network Error \n Please check a network info.\n {err}'
        Dialog.DialogBox(root_view, dialog_txt)
        logging.error(err)
    finally:
        client.close()


def send_cmd_only_for_multi(send_cmd):
    find_ch()
    host = Cons.selected_ch['ip']
    input_port = Cons.selected_ch['port']
    # ic(Cons.selected_ch)
    port = int(0) if Cons.port == '' else int(input_port)
    buf_size = Cons.buf_size

    # 일시적으로 송신 가능 상태가 아니면, False라도 임시로 송신 허용
    was_sending = Cons.data_sending
    if not Cons.data_sending:
        Cons.data_sending = True

    client = socket.socket(AF_INET, SOCK_STREAM)
    try:
        client.settimeout(3)
        if Cons.data_sending:
            client.connect((host, port))
            client.send(bytes(send_cmd))
            reply = client.recv(buf_size)
            # ic(reply.hex())
            Cons.res_log_obj.multi_response(reply.hex())
        else:
            print('Protocol sending was stopped')

    except socket.error as err:
        print(f'network error:{err}')
        dialog_txt = f'Network Error \n Please check a network info.\n {err}'
        # Dialog.DialogBox(root_view, dialog_txt)
        logging.error(err)
    finally:
        client.close()


# Sending Command with hex
def send_cmd_for_uncooled(send_cmd, title, root_view):
    # print('send cmd for uncooled')
    find_ch()
    host = Cons.selected_ch['ip']
    input_port = Cons.selected_ch['port']
    port = int(0) if Cons.port == '' else int(input_port)
    buf_size = Cons.buf_size
    client = socket.socket(AF_INET, SOCK_STREAM)

    try:
        client.settimeout(3)
        # -*- coding: utf-8 -*-
        client.connect((host, port))
        # client.send(bytearray([0xff, 0x00, 0x21, 0x13, 0x00, 0x01, 0x35]))
        # client.send(bytes([0xff, 0x00, 0x21, 0x13, 0x00, 0x01, 0x35]))
        if Cons.data_sending:
            ic('send_cmd_uncooled in Comm', send_cmd)
            client.send(bytes(send_cmd))
            reply = client.recv(buf_size)
            hex_data = binascii.hexlify(reply).decode('utf-8')
            hex_data_14dig = [f'{hex_data[i:i + 14]}' for i in range(0, len(hex_data), 14)]

            if title in Cons.uncooled_query_arrays:
                send_title = title
            else:
                send_title = 'Normal Query'
            # ic('send_cmd_for_uncooled in Comm', reply.hex())
            uncooled_store_response(root_view, send_title, hex_data)

            current_time = datetime.now()
            time_str = current_time.strftime('%Y-%m-%d-%H:%M:%S')
            # hex_with_time = rf'{time_str}: {hex_data_14dig_24space[0]}'
            hex_with_time = fr'{hex_data_14dig[0]} : {time_str}'
            # print(response = {hex_with_time}\n')
            hex_data_14dig[0] = hex_with_time
            for item in hex_data_14dig:
                Cons.response_txt.append(item)
            # print(Cons.response_txt)
            log_pos = Cons.log_txt_fld_info
            log_fld = Res.Response(root_view, log_pos)

        else:
            print('Protocol sending was stopped')

    except socket.error as err:
        print(f'network error:{err}')
        dialog_txt = f'Network Error \n Please check a network info.\n {err}'
        Dialog.DialogBox(root_view, dialog_txt)
        logging.error(err)
    finally:
        client.close()


def send_cmd_for_uncooled_only_cmd(cmd):
    ic(cmd)
    find_ch()
    host = Cons.selected_ch['ip']
    input_port = Cons.selected_ch['port']
    port = int(0) if Cons.port == '' else int(input_port)
    buf_size = Cons.buf_size
    client = socket.socket(AF_INET, SOCK_STREAM)
    try:
        client.settimeout(3)
        client.connect((host, port))
        if Cons.data_sending:
            client.send(bytes(cmd))
            reply = client.recv(buf_size)
            ic('send_cmd_for_uncooled_only_cmd in Comm', reply.hex())
        else:
            print('Protocol sending was stopped')

    except socket.error as err:
        print(f'network error:{err}')
        logging.error(err)

    finally:

        client.close()


def send_cmd_to_ucooled_with_interval(interval: list[float], send_cmds: list[int], cmd_title: list[str], root_view):
    for i, protocol in enumerate(send_cmds):
        # print(i)
        # print(rf'protocol is {protocol}')
        # print(datetime.now())
        if Cons.data_sending:
            if protocol == Cons.capture_hex:
                current_time = datetime.now()
                time_str = current_time.strftime('%Y-%m-%d-%H-%M-%S')
                path = Cons.capture_path['zoom']
                filename = rf'{path}/{time_str}-{i}-{str(cmd_title[i - 1])}.png'
                Mf.capture_image(root_view, filename)
            else:
                if cmd_title[i - 1] in Cons.uncooled_query_arrays:
                    title = cmd_title[i - 1]
                else:
                    title = 'Normal Query'
                if Cons.selected_model == 'Uncooled':
                    # send_cmd_for_uncooled(protocol, title, root_view)
                    send_cmd_for_TTL_uncooled_async(protocol, title, root_view)
                elif Cons.selected_model in ['DRS', 'MiniGimbal', 'Multi']:
                    find_ch()
                    host = Cons.selected_ch['ip']
                    port = int(Cons.selected_ch['port'])
                    send_cmd_for_drs(host, port, protocol, root_view)
                ti.sleep(interval[i])
        else:
            print('Protocol sending was stopped')


# (2024.07.12) Save the response data in 14-digit increments.
def uncooled_store_response(root_view, title, response):
    # res_arrays = []
    # for i in range(0, len(response), 14):
    #     res_data = response[i:i + 14]
    #     res_arrays.append(res_data)
    res_arrays = [response[i:i + 14] for i in range(0, len(response), 14)]
    queries = {
        'Normal Query': (
            'uncooled_normal_q', ['normal']
        ),
        'Zoom Query': (
            'uncooled_zoom_q', ['zoom', 'magnification']
        ),
        'Focus Query': (
            'uncooled_focus_q', ['focus']
        ),
        'Lens Query': (
            'uncooled_lens_q',
            ['af_trigger', 'zoom_spd', 'focus_spd', 'af_mode', 'af_interval', 'fov_position', 'query_mode']
        ),
        'Comm Query': (
            'uncooled_comm_q', ['mode', 'add', 'baud', 'word', 'stop', 'parity', 'protocol']
        ),
        'Image Query': (
            'uncooled_image_q', ['dis', 'dnr', 'flip', 'mirror', 'freeze', 'dzoom', 'dzoom_position']
        ),
        'Sensor Query': (
            'uncooled_sensor_q', ['histogram', 'brightness', 'contrast', 'white_hot', 'pseudo', 'edge']
        ),
        'Cali. Query': (
            'uncooled_cali_q', ['trigger', 'mode', 'interval']
        ),
        'ETC Query': (
            'uncooled_etc_q', ['save', 'cvbs', 'display']
        ),
        'Status Query': (
            'uncooled_status_q', ['boot', 'board_t', 'lens_t', 'sensor_t', 'fan', 'telemetry', 'calibration', 'af']
        ),
        'Version Query': (
            'uncooled_version_q', ['sensor', 'lens', 'main', 'main_y', 'main_m_d', 'osd', 'osd_y', 'osd_m_d']
        ),
        'Encoder Query': (
            'uncooled_encoder_q', ['zoom_max', 'zoom_min', 'focus_max', 'focus_min']
        )
    }
    # print(title)

    if title in queries:
        attr_name, keys = queries[title]
        if hasattr(Cons, attr_name):
            query_dict = getattr(Cons, attr_name)
            for i, key in enumerate(keys):
                if i < len(res_arrays):
                    query_dict[key] = res_arrays[i]
                else:
                    query_dict[key] = []
        else:
            print(f"Error: {attr_name} not found in Cons.")
    else:
        print(f"{title} is not a valid query title.")
        # Cons.uncooled_normal_q['normal'] = res_arrays
        # print(Cons.uncooled_normal_q['normal'])
    # print(Cons.uncooled_normal_q)


# ==================================  2025년 제공을 위해 기능을 OFF ==================================================
#convert_str_with_hex(root_view)


# (2024.07.15) Convert to String from hex Data
def convert_str_with_hex(root_view):
    def extract_hex_value(data_list):
        result = []
        for item in data_list:
            data_chunks = [item[i:i + 2] for i in range(0, len(item), 2)]
            int_value = 0
            if len(data_chunks) > 4:
                msb = data_chunks[4]
                lsb = data_chunks[5]
                hex_str = '0x' + msb + lsb
                int_value = int(hex_str, 16)
                # print(int_value)
            result.append(int_value)
        return result

    # (2024.07.16) Get the System Info. Table Data From Constant
    zoom_msb_lsb = [Cons.uncooled_zoom_q['zoom'], Cons.uncooled_lens_q['zoom_spd'],
                    Cons.uncooled_image_q['dzoom'], Cons.uncooled_image_q['dzoom_position']]
    focus_msb_lsb = [Cons.uncooled_focus_q['focus'], Cons.uncooled_lens_q['focus_spd'], Cons.uncooled_lens_q['af_mode']]
    fov_msb_lsb = [Cons.uncooled_lens_q['fov_position']]

    zoom_str = extract_hex_value(zoom_msb_lsb)
    focus_str = extract_hex_value(focus_msb_lsb)
    fov_str = extract_hex_value(fov_msb_lsb)
    Cons.zoom_msb_lsb = zoom_str
    Cons.focus_msb_lsb = focus_str
    Cons.fov_msb_lsb = fov_str
    SysInfo.SysInfo(root_view, Cons.sys_info_tab)


# (2024.07.19): Open socket and send cmd to nyx
def send_cmd_to_nyx(root, cmd):
    find_ch()
    host = Cons.selected_ch['ip']
    input_port = Cons.selected_ch['port']
    send_port = int(0) if Cons.port == '' else int(input_port)
    buf_size = Cons.buf_size
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.connect((host, send_port))
        s.sendall(cmd.encode('utf-8'))
        response = s.recv(buf_size)

    return response
    # current_time = datetime.now()
    # time_str = current_time.strftime('%Y-%m-%d-%H:%M:%S')
    #
    # print(f"Received response: {response.decode('utf-8')}")
    # response_with_time = fr'{time_str} : {response.decode('utf-8')}'
    # Cons.response_txt.append(response_with_time)
    # log_pos = Cons.log_txt_fld_info
    # log_fld = Res.Response(root, log_pos)


def send_cmd_to_nyx_without_root(cmd):
    find_ch()
    host = Cons.selected_ch['ip']
    input_port = Cons.selected_ch['port']
    send_port = int(0) if Cons.port == '' else int(input_port)
    buf_size = Cons.buf_size
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.connect((host, send_port))
        s.sendall(cmd.encode('utf-8'))
        response = s.recv(1024)

        current_time = datetime.now()
        time_str = current_time.strftime('%Y-%m-%d-%H:%M:%S')

        print(f"Received response: {response.decode('utf-8')}")
        response_with_time = fr'{time_str} : {response.decode('utf-8')}'
        Cons.response_txt.append(response_with_time)


# 2025.05.22: Added to press registerBTN in main
def click_register_button(app_instance):
    try:
        app_instance.register_btn.invoke()
        print("register_btn was pressed.")
    except AttributeError:
        print("register_btn was not defined")


# 2025.05.22: Added to wait until NYX is accessible
def wait_for_nyx_ready(host, port, retries=10, delay=5):
    for _ in range(retries):
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as test_socket:
                test_socket.settimeout(3)
                test_socket.connect((host, port))
                print('NYX was connected.')
                return True
        except:
            print('NYX is booting retry')
            ti.sleep(delay)
    return False


# 2024.07.24: Open socket and store to Constant after send cmd with interval
# 2025.05.22: Retry Connect to NYX after send Reboot CMD for NYX
# 2025.11.24 Async to NYX Script and log text file was applied
def send_cmd_to_nyx_with_interval(app, root, titles, cmds, intervals_sec, response_file_name):
    send_cmds = [create_form(cmd) for cmd in cmds]
    results = []

    for i, s_cmd in enumerate(send_cmds):
        if not Cons.data_sending:
            print('Data sending stopped by user. Exiting loop.')
            return

        ic(rf'{i + 1} is {s_cmd}')
        current_time = datetime.now()
        time_str = current_time.strftime('%Y-%m-%d-%H-%M-%S')

        if titles[i] == Cons.capture_title:
            path = Cons.capture_path['zoom']
            filename = rf'{path}/{str(titles[i - 3])}-{time_str}-{i}.png'
            Mf.capture_image(root, filename)
        elif s_cmd.startswith('NYX.SET#syst_exec=reboot'):
            find_ch()
            host = Cons.selected_ch['ip']
            input_port = Cons.selected_ch['port']
            send_port = int(0) if Cons.port == '' else int(input_port)
            print('reboot')
            # s.sendall(s_cmd.encode('utf-8'))
            Async.async_send(fn=lambda cmd=s_cmd: send_cmd_to_nyx(root, cmd), title=titles[i],
                                   root_view=root, log_name=response_file_name)
            ti.sleep(10)
            if not Cons.data_sending:
                print('Stopped during reboot wait.')
                return

            if wait_for_nyx_ready(host, send_port):
                s = create_socket()
                ti.sleep(3)

                if not Cons.data_sending:
                    print('Stopped during reboot wait.')
                    return

                click_register_button(app)
                ti.sleep(100)

                if not Cons.data_sending:
                    print('Stopped during reboot wait.')
                    return

            else:
                print("NYX reconnect fail")
                return
        else:
            Async.async_send(fn=lambda cmd=s_cmd: send_cmd_to_nyx(root, cmd), title=titles[i],
                                   root_view=root, log_name=response_file_name )
            ti.sleep(intervals_sec[i])

            if not Cons.data_sending:
                print('Stopped during reboot wait.')
                return


# 2024.07.19): Calculate LRC for NYX
def calc_lrc(send_form):
    lrc_key = 0x40
    lrc = 0

    for char in send_form:
        # convert each character of string to unicode int
        lrc += ord(char)

    lrc = (lrc ^ 0xFF) + 0x01
    lrc_hi = lrc_key + ((lrc >> 4) & 0x0F)
    lrc_lo = lrc_key + (lrc & 0x0F)

    return chr(lrc_hi), chr(lrc_lo)


# (2024.07.19): Create a Protocol form for NYX
def create_form(cmd):
    lrc_hi, lrc_lo = calc_lrc(cmd)
    cmd += lrc_hi + lrc_lo + '\n'
    return cmd


# (2024.07.19): Protocol send function for NYX
def nyx_cmd_to_convert(event, root):
    tree = event.widget
    # get id of select item
    selected_item = tree.selection()[0]
    # get item of selected id
    item = tree.item(selected_item)
    values = item['values'][1]
    form = create_form(values)

    return form


# (2024.7.22): Protocol send with command for NYX
def send_data_with_cmd_for_nyx_ptz(root, cmd):
    value = cmd
    print(f'value = {value}')
    form = create_form(value)
    send_cmd_to_nyx(root, form)


def send_data_with_cmd_for_nyx_ptz_without_root(cmd):
    value = cmd
    print(f'value = {value}')
    form = create_form(value)
    send_cmd_to_nyx_without_root(form)


# (2024.07.30): query for information
def send_data_with_cmd_for_info(root, cmds):
    host = Cons.host_ip
    input_port = Cons.port
    send_port = int(0) if Cons.port == '' else int(input_port)
    buf_size = Cons.buf_size

    responses = []
    for cmd in cmds:
        form = create_form(cmd)
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.connect((host, send_port))
            s.sendall(form.encode('utf-8'))
            response = s.recv(1024)
            res_decode = response.decode('utf-8')
            start_digi = len(cmd)
            info_data = res_decode[start_digi:-3]
            responses.append(info_data)
    Cons.cooled_lens_pos_spd = responses


# (2024.08.13): Added send cmd function for DRS
def send_cmd_for_drs(host, port, send_cmd, root_view):
    # print('send_cmd_for_drs')
    buf_size = Cons.buf_size
    client = socket.socket(AF_INET, SOCK_STREAM)
    try:
        client.settimeout(3)
        client.connect((host, port))
        client.send(bytes(send_cmd))
        reply = client.recv(buf_size)
        # print(reply)

        if Cons.selected_model == 'DRS':
            # 데이터 파싱 함수 호출
            parse_drs_reply(reply)
        else:
            return

        # 로그 창에 응답 표시
        log_pos = Cons.log_txt_fld_info
        log_fld = Res.Response(root_view, log_pos)
    except socket.error as err:
        handle_network_error(err, root_view)
    finally:
        client.close()


def parse_drs_reply(reply):
    lsb = reply[::2]
    msb = reply[1::2]
    binary_pairs = [format(m, '08b') + format(l, '08b') for m, l in zip(msb, lsb)]

    handlers = {
        1: DR_r.check_data01,
        2: DR_r.check_data02,
        3: lambda binary: update_drs_response('roi_x_start_x_pos', DR_r.convert_to_deci(binary)),
        4: lambda binary: update_drs_response('roi_x_start_y_pos', DR_r.convert_to_deci(binary)),
        5: DR_r.check_data05,
        6: DR_r.check_data06,
        7: lambda binary: update_drs_response('contrast', DR_r.convert_to_deci(binary)),
        8: lambda binary: update_drs_response('brightness', DR_r.convert_to_deci(binary)),
        9: DR_r.check_data09,
        10: lambda binary: update_drs_response('serial', format(DR_r.convert_to_deci(binary), 'x')),
        11: lambda binary: update_drs_response('shutter_temp', (DR_r.convert_to_deci(binary) - 27300) / 100),
        12: lambda binary: update_drs_response('roi_x_threshold_temp', format(DR_r.convert_to_deci(binary), 'x')),
        13: lambda binary: update_drs_response('center_temp', (DR_r.convert_to_deci(binary) / 10)),
        14: DR_r.check_data14,
        15: lambda binary: update_drs_response('zoom_pos', DR_r.convert_to_deci(binary)),
        16: lambda binary: update_drs_response('focus_pos', DR_r.convert_to_deci(binary)),
        17: DR_r.check_data17,
        18: DR_r.check_data18,
        19: lambda binary: update_drs_response('frame_min_temp', (DR_r.convert_to_deci(binary) / 10)),
        20: lambda binary: update_drs_response('frame_max_temp', (DR_r.convert_to_deci(binary) / 10)),
        21: lambda binary: update_drs_response('frame_average_temp', (DR_r.convert_to_deci(binary) / 10)),
        22: lambda binary: update_drs_response('roi0_min', (DR_r.convert_to_deci(binary) / 10)),
        23: lambda binary: update_drs_response('roi0_max', (DR_r.convert_to_deci(binary) / 10)),
        24: lambda binary: update_drs_response('roi1_min', (DR_r.convert_to_deci(binary) / 10)),
        25: lambda binary: update_drs_response('roi1_max', (DR_r.convert_to_deci(binary) / 10)),
        26: lambda binary: update_drs_response('roi2_min', (DR_r.convert_to_deci(binary) / 10)),
        27: lambda binary: update_drs_response('roi2_max', (DR_r.convert_to_deci(binary) / 10)),
        28: lambda binary: update_drs_response('roi3_min', (DR_r.convert_to_deci(binary) / 10)),
        29: lambda binary: update_drs_response('roi3_max', (DR_r.convert_to_deci(binary) / 10)),
        30: lambda binary: update_drs_response('roi4_min', (DR_r.convert_to_deci(binary) / 10)),
        31: lambda binary: update_drs_response('roi4_max', (DR_r.convert_to_deci(binary) / 10)),
        32: lambda binary: update_drs_response('roi5_min', (DR_r.convert_to_deci(binary) / 10)),
        33: lambda binary: update_drs_response('roi5_max', (DR_r.convert_to_deci(binary) / 10)),
        34: lambda binary: update_drs_response('roi6_min', (DR_r.convert_to_deci(binary) / 10)),
        35: lambda binary: update_drs_response('roi6_max', (DR_r.convert_to_deci(binary) / 10)),
        36: lambda binary: update_drs_response('roi7_min', (DR_r.convert_to_deci(binary) / 10)),
        37: lambda binary: update_drs_response('roi7_max', (DR_r.convert_to_deci(binary) / 10)),
        38: lambda binary: update_drs_response('roi8_min', (DR_r.convert_to_deci(binary) / 10)),
        39: lambda binary: update_drs_response('roi8_max', (DR_r.convert_to_deci(binary) / 10)),
        40: lambda binary: update_drs_response('roi9_min', (DR_r.convert_to_deci(binary) / 10)),
        41: lambda binary: update_drs_response('roi9_max', (DR_r.convert_to_deci(binary) / 10)),
        42: lambda binary: update_drs_response('colorbar_min', (DR_r.convert_to_deci(binary) / 10)),
        43: lambda binary: update_drs_response('colorbar_max', (DR_r.convert_to_deci(binary) / 10)),
        44: lambda binary: update_drs_response('emissivity', (DR_r.convert_to_deci(binary) / 100)),
        45: lambda binary: update_drs_response('user_temp_offset', (DR_r.convert_to_deci(binary) / 100)),
        46: DR_r.check_data46,
        47: DR_r.check_data47,
        48: lambda binary: update_drs_response('roi_x_end_pos', DR_r.convert_to_deci(binary)),
        49: lambda binary: update_drs_response('check_data', DR_r.convert_to_deci(binary)),

    }

    for i, binary_pair in enumerate(binary_pairs):
        if i in handlers:
            handlers[i](binary_pair)


def update_drs_response(key, value):
    Cons.drs_response[key] = value


def handle_network_error(err, root_view):
    print(f'network error: {err}')
    dialog_txt = f'Network Error \n Please check a network info.\n {err}'
    Dialog.DialogBox(root_view, dialog_txt)
    logging.error(err)


# (2024.09.24): PTZ func for FT
# (2024.10.14): Applied Http Digest
@staticmethod
def fine_tree_send_cgi(url, params):
    find_ch()
    # 2025.04.29: Was Applied Port No (FineTree needs Web Port No for Controlling)
    sel_ip = Cons.selected_ch['ip'] + f':{Cons.selected_ch['port']}'
    print(rf'sel ip is {sel_ip}')

    base_url = rf'http://{sel_ip}{url}?'

    if len(params) == 1:
        username = Cons.selected_ch['id']
        password = Cons.selected_ch['pw']

        params_encoded = urllib.parse.urlencode(params)
        print(f"Sending request to: {base_url} with params: {params_encoded}")
        full_url = f'http://{sel_ip}{url}?{params_encoded}'
        try:
            response = requests.get(base_url, params=params_encoded, auth=HTTPBasicAuth(username, password), timeout=5)
            if response.status_code == 200:
                print("Request was successful!")
                print(response.text)
            else:
                print(f"Failed to send request. Status code: {response.status_code}")
        except requests.exceptions.Timeout:
            print("The request timed out.")
        except requests.exceptions.ConnectionError as e:
            print(f"Connection error occurred: {e}")
        except requests.exceptions.RequestException as e:
            print(f"An error occurred: {e}")
    else:
        if 'group' in params.keys():
            get_url = rf'http://{sel_ip}{url}?group={params['group']}&app=get'
            print(get_url)
            set_url = rf'{url}?group={params["group"]}&app={params["app"]}&{list(params.keys())[2]}={list(params.values())[2]}'
            print(set_url)
        elif any(x in list(params.keys())[1] for x in ['restart', 'reset', 'default']):
            print('any')
            get_url = rf'http://{sel_ip}{url}?app=get'
            print(get_url)
            set_url = rf'{url}?app={params["app"]}'
            print(set_url)
        else:
            print('else')
            get_url = rf'http://{sel_ip}{url}?app=get'
            print(rf'get url is {get_url}')
            set_url = rf'{url}?app={params["app"]}&{list(params.keys())[1]}={list(params.values())[1]}'
            print(set_url)

        try:
            get_auth = get_auth_parameters(sel_ip, get_url)
            print(get_auth)
            response = send_authenticated_request(get_auth, set_url, method='GET')
            print(response)
            print("Response status code:", response.status_code)
            print("Response body:", response.text)
        except Exception as e:
            print("Error:", str(e))


# (2024.11.14): Sending a PTZ CMD to Finetree by DRS
def send_cmd_to_Finetree(url, params):
    Cons.selected_model = 'FineTree'
    find_ch()
    sel_ip = Cons.selected_ch['ip']
    print(sel_ip)
    base_url = rf'http://{sel_ip}:{Cons.selected_ch['port']}{url}?'
    username = Cons.selected_ch['id']
    password = Cons.selected_ch['pw']

    params_encoded = urllib.parse.urlencode(params)
    print(f"Sending request to: {base_url} with params: {params_encoded}")
    full_url = f'http://{sel_ip}{url}?{params_encoded}'
    try:
        response = requests.get(base_url, params=params_encoded, auth=HTTPBasicAuth(username, password), timeout=5)
        if response.status_code == 200:
            print("Request was successful!")
            print(response.text)
        else:
            print(f"Failed to send request. Status code: {response.status_code}")
    except requests.exceptions.Timeout:
        print("The request timed out.")
    except requests.exceptions.ConnectionError as e:
        print(f"Connection error occurred: {e}")
    except requests.exceptions.RequestException as e:
        print(f"An error occurred: {e}")

    Cons.selected_model = 'DRS'
    find_ch()
    print(rf'last model is {Cons.selected_model}')


# (2024.09.25) Find selected model
# 2025.06.30: Added a PT Driver
@staticmethod
def find_ch():
    model = Cons.selected_model
    model_arrays = [Cons.ch1_rtsp_info, Cons.ch2_rtsp_info,
                    Cons.ch3_rtsp_info, Cons.ch4_rtsp_info,
                    Cons.pt_drv_info]
    for i, ch in enumerate(model_arrays):
        if model == ch['model']:
            Cons.selected_ch = model_arrays[i]


# ========================================= Regarding Create Only One Socket ===========================================
# (2024.12.06) Create Only One Socket
def create_socket() -> socket.socket:
    client_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    int_port = int(Cons.selected_ch['port'])
    client_socket.connect((Cons.selected_ch['ip'], int_port))

    return client_socket


def send_to_mini(sock: socket.socket, cmd):
    try:
        sock.send(bytes(cmd))
        # reply = sock.recv(78)
        # print(reply)
        #
        # return reply
    except Exception as e:
        print("error while sending cmd to mini :", str(e))


def close_socket(sock: socket.socket):
    print('socket was closed')
    sock.close()


def send_to_mini_with_interval(sock: socket.socket, titles, cmds, intervals):
    for i, protocol in enumerate(cmds):
        if Cons.data_sending:
            # print(rf'{datetime.now()}: {titles[i]}')
            try:
                sock.send(bytes(protocol))
                ti.sleep(intervals[i])
                reply = sock.recv(39)
                # print(reply)
            except Exception as e:
                print("error while sending cmd to mini :", str(e))
        else:
            return


# (2024.12.10) Save a mini gimbal response to text file
def save_res_from_miniG_Text(response):
    res_arrs = []
    # parser = [int.from_bytes(reply[i:i+1], byteorder='big') for i in range(0, len(reply))]
    hex_value = [f'{bytes:02x}' for bytes in response]
    # print(hex_value)
    for i, value in enumerate(hex_value):
        if i != 38 and i != 77:
            if value == 'ff' and hex_value[i + 1] == '01' and hex_value[i + 2] == '23':
                if len(hex_value) == 39:
                    res_arrs = hex_value[i:i + 39]

                    # print(res_arrs)
                    update_res_to_cons(res_arrs)
                    file_path = rf'Log/mini_gimbal.txt'
                    try:
                        with open(file_path, "r") as file:
                            existing_data = file.readlines()
                    except FileNotFoundError:
                        existing_data = []

                    with open(file_path, "w") as file:
                        file.write(rf'{datetime.now()}: ' + ' '.join(res_arrs) + '\n')
                        file.writelines(existing_data)


# (2024.12.11) Save a mini gimbal response to csv file
# I don't know why csv file was not applied when received a response
def save_res_from_miniG_CSV(response):
    column_title = Cons.miniG_res_row
    # res_arrs = []
    hex_value = [f'{bytes:02x}' for bytes in response]
    file_path = rf'Log/mini_gimbal.xlsx'

    with file_lock:
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Mini Gimbal Response'

            # Add the header
            for i, title in enumerate(column_title, start=1):
                ws.cell(row=1, column=i + 5, value=title)

        for i, value in enumerate(hex_value):
            # print(hex_value)
            if i != 38 and i != 77:
                if value == 'ff' and hex_value[i + 1] == '01' and hex_value[i + 2] == '23':
                    if len(hex_value) >= 39:
                        res_arrs = hex_value[i:i + 39]
                        print(rf'{datetime.now()}: {res_arrs}')

                        next_row = ws.max_row + 1
                        for col_index, data in enumerate(res_arrs):
                            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')
                            ws.cell(row=next_row, column=1, value=current_time)
                            ws.cell(row=next_row, column=col_index + 2, value=data)
                            wb.save(f'{file_path}')


def update_res_to_cons(res_arrs):
    Cons.miniG_res_payload.update({
        'roll_en_hi': res_arrs[4],
        'roll_en_lo': res_arrs[5],
        'pitch_en_hi': res_arrs[6],
        'pitch_en_lo': res_arrs[7],
        'yaw_en_h': res_arrs[8],
        'yaw_en_l': res_arrs[9],
        'cam_status': res_arrs[10],
        'fan_heater_sta': res_arrs[11],
        'motor_bd': res_arrs[12],
        'temp': res_arrs[13],
        'drift_offset_h': res_arrs[14],
        'drift_offset_l': res_arrs[15],
        'stabilizer_mode': res_arrs[16],
        '17': res_arrs[17],
        'eo_dzoom': res_arrs[18],
        'eo_wdr': res_arrs[19],
        'eo_blc': res_arrs[20],
        'eo_dis': res_arrs[21],
        'eo_dn': res_arrs[22],
        '23': res_arrs[23],
        '24': res_arrs[24],
        'eo_defog': res_arrs[25],
        'eo_op_zoom_h': res_arrs[26],
        'eo_op_zoom_l': res_arrs[27],
        'eo_d_zoom': res_arrs[28],
        'eo_bri': res_arrs[29],
        'eo_sharp': res_arrs[30],
        'ir_dzoom': res_arrs[31],
        'ir_pale': res_arrs[32],
        'ir_agc_mode': res_arrs[33],
        'ir_dde': res_arrs[34],
        '35': res_arrs[35],
        'fw_h': res_arrs[36],
        'fw_l': res_arrs[37],
    })
    print(rf'{datetime.now()}: {Cons.miniG_res_payload}')
    roll = convert_angle(res_arrs[4], res_arrs[5])
    pitch = convert_angle(res_arrs[6], res_arrs[7])
    yaw = convert_angle(res_arrs[8], res_arrs[9])
    print(rf'roll is {roll}')
    print(rf'pitch is {pitch}')
    print(rf'yaw is {yaw}')


# 2024.12.30: added convert an angle from hex data
def convert_angle(high, low) -> float:
    high = int(high, 16)
    low = int(low, 16)
    raw = (high << 8) + low
    angle = 0.0
    if raw & 0x8000:
        deci_raw = 0x10000 - raw
        angle = (deci_raw * 180) / -32767
    else:
        angle = (raw * 180) / 32767
    return angle


# Test Code
def run():
    parms_array = [{'move': 'up'}, {'move': 'down'}, {'move': 'left'}, {'move': 'right'},
                   {'move': 'upright'}, {'move': 'downleft'}, {'move': 'upleft'}, {'move': 'downright'}]
    for parms in parms_array:
        fine_tree_send_cgi(parms)
        ti.sleep(1)
        parm_stop = {'move': 'stop'}
        fine_tree_send_cgi(parm_stop)
        ti.sleep(1)


# ================================================= Check a MD5 ========================================================
def md5_hexdigest(data):
    return hashlib.md5(data.encode('utf-8')).hexdigest()


def get_auth_parameters(host_ip, get_url):
    # Step 1: Get the authentication parameters from the server
    url = rf'http://{host_ip}/setup/video/image.php?group=basic&app=get'
    response = requests.get(url)

    if response.status_code == 401:
        # Parsing the WWW-Authenticate header
        www_authenticate = response.headers['www-authenticate']
        print(www_authenticate)
        params = {}

        for param in www_authenticate.split(','):
            key, value = param.strip().split('=')
            params[key] = value.strip('"')

        return params
    else:
        raise Exception("Failed to get authentication parameters")


def create_digest_response(realm, nonce, qop, nc, cnonce, method, uri):
    # Step 2: Calculate the HA1, HA2, and response
    ha1 = md5_hexdigest(f"{Cons.selected_ch['id']}:{realm}:{Cons.selected_ch['pw']}")
    ha2 = md5_hexdigest(f"{method}:{uri}")
    response = md5_hexdigest(f"{ha1}:{nonce}:{nc}:{cnonce}:{qop}:{ha2}")

    return response


def send_authenticated_request(params, uri, method='GET', additional_params=None):
    realm = params['Digest realm']
    nonce = params['nonce']
    qop = params['qop']

    # Create a unique client nonce
    cnonce = uuid.uuid4().hex
    nc = '00000001'  # Nonce count (increment for each request with the same nonce)

    # Create the digest response
    response = create_digest_response(realm, nonce, qop, nc, cnonce, method, uri)

    # Construct the Authorization header
    auth_header = (
        f'Digest realm="{realm}",'
        f' qop="{qop}",'
        f' nonce="{nonce}",'
        f' opaque="{params["opaque"]}",'
        f' username="{Cons.selected_ch['id']}",'
        f' algorithm="MD5",'
        f' nc={nc},'
        f' cnonce="{cnonce}",'
        f' response="{response}",'
        f' uri="{uri}"'
    )

    # Step 3: Send the authenticated request
    full_url = f"http://{Cons.selected_ch['ip']}:{Cons.selected_ch['port']}/{uri}"
    if additional_params:
        full_url += '&' + '&'.join(f"{key}={value}" for key, value in additional_params.items())

    headers = {
        'Authorization': auth_header,
        'Connection': 'close'
    }

    response = requests.request(method, full_url, headers=headers)

    return response


# 2025.11.17: threading function was applied

def send_cmd_for_TTL_uncooled_async(send_cmd_bytes, title=None, root_view=None):
    """
    - send_cmd_bytes: list[int], str(hex), bytes 등 아무거나 받아도 됨
    - UI 스레드는 바로 리턴되고, 통신은 백그라운드 스레드에서 수행
    """
    # 0) 먼저 payload를 bytes로 통일
    try:
        payload = Cal.to_bytes_payload(send_cmd_bytes)  # <-- 여기서만 변환
    except Exception as e:
        logging.exception(e)
        raise TypeError("send_cmd_bytes must be bytes") from e

    def worker():
        try:
            # 1) 여기서는 "동기 함수"를 호출해야 함 (중요!!)
            rx = send_cmd_for_TTL_uncooled(payload, title, root_view)
            #    ↑ 이 함수가 실제로 TLS 연결 + Digest + fwtransparent + recv 다 하는 기존 함수
        except Exception as e:
            logging.exception("uncooled async send error: %s", e)
            if root_view is not None:
                # UI 업데이트는 메인 스레드에서
                root_view.after(0, lambda: logging.error(f"[ERR] {title}: {e}"))
            return

        # 2) 응답을 UI에 반영 (역시 메인 스레드에서만)
        if root_view is not None:
            def _update_ui():
                hex_data = binascii.hexlify(rx).decode("utf-8")
                logging.info("uncooled async handled(%s): %s", title, hex_data)
                # 필요하면 여기서 로그창/Treeview 업데이트
                # ex) uncooled_store_response(root_view, title, hex_data)

            root_view.after(0, _update_ui)

    t = threading.Thread(target=worker, daemon=True)
    t.start()


# 2025.11.16: TTL Auth has been applied applied to Seyeon
def send_cmd_for_TTL_uncooled(send_cmd, title, root_view):
    find_ch()
    user_info = Cons.selected_ch

    # 0) 안전장치
    if not isinstance(send_cmd, (bytes, bytearray)):
        raise TypeError("send_cmd_bytes must be bytes")
    payload = bytes(send_cmd) + b"\r\n"  # ★ CRLF 필수

    # 1) Digest 재료와 쿠키 준비
    chal = ttl._get_admin_challenge(user_info['ip'])
    cookie = ttl._enc_cookie(user_info['id'], user_info['pw'])

    # 2) fwtransparent 요청라인 만들기
    uri = (f"/cgi-bin/fwtransparent.cgi"
           f"?BaudRate={Cons.FW_BAUD}&DataBit={Cons.FW_DBITS}&StopBit={Cons.FW_SBITS}"
           f"&ParityBit={Cons.FW_PARITY}&Node={Cons.FW_NODE}&FwCgiVer=0x0001")
    authz = ttl._build_digest_md5(user_info['id'], user_info['pw'], "GET", uri, chal)

    # 3) TLS 연결 후 GET 헤더 송신(응답은 기다리지 않고 터널 용도로 사용)
    s = ttl._tls_connect(user_info['ip'], user_info['port'])
    try:
        req = (f"GET {uri} HTTP/1.0\r\n"
               f"Host: {user_info['ip']}\r\n"
               f"Authorization: {authz}\r\n"
               f"Cookie: -goahead-session-={cookie}\r\n"
               "Accept: */*\r\nUser-Agent: py-client\r\n"
               "Connection: keep-alive\r\n\r\n").encode()
        s.sendall(req)

        # 4) 터널 준비 시간
        time.sleep(Cons.PRE_SEND_DELAY)

        # 5) 명령 송신
        s.sendall(payload)

        # 6) 응답 수신(최대 RECV_WAIT_SEC)
        end_by = time.time() + Cons.RECV_WAIT_SEC
        buf = bytearray()
        while time.time() < end_by:
            try:
                chunk = s.recv(Cons.READ_MAX_BYTES)
                if not chunk:
                    break
                buf += chunk
            except socket.timeout:
                break

        rx = bytes(buf)
        # ★ 로그/가공 예시 (필요 시 기존 UI/저장 함수 호출)
        logging.info("uncooled tx=%s", binascii.hexlify(payload).decode())
        logging.info("uncooled rx(%d)=%s", len(rx), binascii.hexlify(rx).decode())
        return rx

    finally:
        try:
            s.shutdown(socket.SHUT_WR)
        except Exception:
            pass
        s.close()


# 2025.11.24 Control a Log Text File
def log_control(text, log_dir="Log", file_name=None):
    log_dir = Cons.log_path.rstrip('/\\')  # 끝의 / 또는 \ 제거
    os.makedirs(log_dir, exist_ok=True)

    if file_name is None:
        time_str = datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
        file_name = f"log_{time_str}.txt"

    full_path = os.path.join(log_dir, file_name)
    try:
        with open(full_path, 'a') as f:
            f.write(text + '\n')
    except Exception as e:
        logging.error("Log file write error(%s): %s", full_path, e)
        return None
