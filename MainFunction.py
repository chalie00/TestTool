import threading

import openpyxl
import tkinter
import string
import mss
import mss.tools
import time as ti
import tkinter.font as tkfont

import Dialog
import Communication as Comm
import Constant as Cons
import Table as tb
import KeyBind as Kb

from tkinter import *
from tkinter import ttk
from ttkwidgets import CheckboxTreeview
from screeninfo import get_monitors
from datetime import datetime
from icecream import ic

from Communication import send_data_for_nyx


# Set an element (label, text field, button) as specified position and size
def make_element(x, y, h, w, element, *args, **kwargs):
    f = Frame(height=h, width=w)
    f.pack_propagate(0)  # don't shrink
    f.place(x=x, y=y)
    if element.lower() == 'label':
        label = Label(f, *args, **kwargs)
        label.pack(fill=BOTH, expand=1)
        return label
    elif element.lower() == 'entry':
        text_field = Entry(f, *args, **kwargs)
        text_field.pack(fill=BOTH, expand=1)
        return text_field
    elif element.lower() == 'button':
        btn = Button(f, *args, **kwargs)
        btn.pack(fill=BOTH, expand=1)
        return btn


def fix_selection_tags(event):
    tree = event.widget
    selected_iids = tree.selection()
    for iid in selected_iids:
        # 현재 tag 다시 설정해서 선택 상태에서도 유지
        tags = tree.item(iid, 'tags')
        tree.item(iid, tags=tags)


# Set Command Table
# 2025.06.25: Added a CTEC Protocol of Multi Sensor
# (2024.02.15) CheckBox Change a treeview to CheckTreeView  and event is changed to
#  Double-Button-1, CheckBox is controlled with tag
def make_table(root: tkinter, column_num: int, width: int, column_title: [string], x: int, y: int,
               cmd_data: []) -> CheckboxTreeview:
    # Create a table, column is the name of column,
    # The display column shows the order in which the table is executed.

    column = [i + 1 for i in range(column_num)]
    dis_column = [str(n) for n in column]
    tv = CheckboxTreeview(root, height=29, columns=dis_column, displaycolumns=dis_column)

    # 2025.06.13: Pgup/Pgdn Direction key prevent
    tv.bind("<Prior>", lambda e: Kb.pressed_kbd_direction(e))
    tv.bind("<Next>", lambda e: Kb.pressed_kbd_direction(e))
    tv.bind("<KeyPress-Up>", lambda e: Kb.pressed_kbd_direction(e))
    tv.bind("<KeyPress-Down>", lambda e: Kb.pressed_kbd_direction(e))
    tv.bind("<KeyPress-Left>", lambda e: Kb.pressed_kbd_direction(e))
    tv.bind("<KeyPress-Right>", lambda e: Kb.pressed_kbd_direction(e))

    # set the treeview scroll
    vsb = ttk.Scrollbar(root, orient='vertical', command=tv.yview)
    vsb.place(x=width * column_num + 105 + Cons.cam1_resolution['w'], y=y, height=Cons.tree_view_size['h'] - 30)
    # tv = tkinter.ttk.Treeview(root, columns=column, display columns=dis_column)
    # Treeview의 width, height 글자 수로 정해 진다.
    # tv.configure(height=len(Cons.command_array) + 1)
    tv.place(x=x, y=y)
    tv.configure(yscrollcommand=vsb.set)

    # Set the Table No Tab
    tv.column('#0', width=70, anchor='center', stretch='yes')
    tv.heading('#0', text='No', anchor='center')

    # Create each column (name, width, anchor)
    for i in range(column_num):
        tv.column(dis_column[i], width=width, anchor='center')
        tv.heading(dis_column[i], text=column_title[i], anchor='center')

    default_font = tkfont.nametofont("TkTextFont")
    bold_font = default_font.copy()
    bold_font.configure(weight="bold")

    tv.tag_configure('cmd_tag', background='lightblue')
    tv.tag_configure('query_tag', background='yellow')
    tv.tag_configure('bold_text', font=bold_font)

    # Insert the command data in Table
    for i, row_values in enumerate(cmd_data):
        tags = ['unchecked']  # 기본 태그

        # 태그 조건 누적
        if any(cell == 'CMD List' for cell in row_values):
            tags.append('cmd_tag')
        if any('Query' in str(cell) for cell in row_values):
            tags.append('query_tag')
            tags.append('bold_text')

        # 단 한 번만 insert, 태그는 누적해서
        tv.insert('', 'end', text=i + 1, values=row_values, iid=f'{i}번', tags=tags)

    # 더블 클릭 바인딩은 루프 밖에서 1번만!
    tv.bind('<Double-Button-1>', lambda event, root_view=root, tree=tv: clicked_table_element(event, root_view, tv))
    tv.bind('<<TreeviewSelect>>', fix_selection_tags)
    return tv


def convert_str_to_hex(hex_str_arr) -> []:
    hex_array = []
    for i in range(0, len(hex_str_arr), 2):
        hex_str = '0x' + hex_str_arr[i:i + 2]
        hex_int = int(hex_str, 16)
        hex_array.append(hex_int)
    print('hex here')
    return hex_array


def select_item(event, root_view) -> []:
    hex_array = []
    tree = event.widget
    # selection = [tree.item(item)["text"] for item in tree.selection()]
    selection = [tree.item(item)['values'][1] for item in tree.selection()]
    for i in range(0, len(selection[0]), 2):
        hex_str = '0x' + selection[0][i:i + 2]
        hex_int = int(hex_str, 16)
        hex_array.append(hex_int)

    return hex_array


# Check Whether Interval Button Active
def check_interval_active():
    Cons.interval_button.config(state='normal')
    cmd_title_count = len(Cons.script_cmd_titles)
    interval_count = len(Cons.script_itv_arrs)
    if cmd_title_count > interval_count:
        Cons.interval_button.configure(state='normal')
    else:
        Cons.interval_button.configure(state='disabled')


# Called when a table element was clicked
def clicked_table_element(event, root_view, tv):
    Comm.find_ch()
    # print(Cons.selected_model)
    host = Cons.selected_ch['ip']
    input_port = Cons.selected_ch['port']
    port = int(0) if Cons.port == '' else int(input_port)
    ic(rf'ip:{host}, port:{input_port}, port:{port}')

    iden = tv.identify_row(event.y)
    tags = tv.item(iden, 'tags')
    item = tv.item(iden)
    # item['values'][0] is title that user was clicked
    # item['values'][1] is command
    selected_item = item['values']
    title = selected_item[0]
    value = selected_item[1]

    if not host or not port:
        print("Invalid command")
        show_network_dialog(root_view, 'Please enter a network info.')
        return

    if Cons.script_toggle_flag:
        # handle_script_mode(event, iden, title, root_view)
        handle_script_mode(event, iden, selected_item, root_view)
    else:
        handle_normal_mode(event, tags, iden, title, root_view, tv, host, port)


# 2024.07.04: Operating script mode
# (2024.07.25): NYX added to script mode
# (2024.08.14): DRS Added to script mode
# (2024.10.18): FineTree Added to script mode
def handle_script_mode(event, iden, value, root_view):
    Cons.data_sending = True
    if Cons.selected_model in ['Uncooled', 'DRS', 'MiniGimbal', 'Multi']:
        hex_value = select_item(event, root_view)
        Cons.script_cmd_arrs.append(hex_value)
        Cons.script_cmd_titles.append(value[0])
        print(Cons.script_cmd_titles)
        print(hex_value)

        gene_interval_arrays(value, root_view)
        check_interval_active()

    elif Cons.selected_model == 'NYX Series':
        Cons.script_cmd_titles.append(value[0])
        # converted_cmd = Comm.create_form(value[1])
        Cons.script_cmd_arrs.append(value[1])

        gene_interval_arrays(value, root_view)
        check_interval_active()

    elif Cons.selected_model == 'FineTree':
        index = int(iden.split('번')[0])
        # fine_tree_cmd_data = (para_arr, value_arr, url)
        items = Cons.fine_tree_cmd_data[index]
        # print(rf'items is {items}')
        Cons.script_cmd_titles.append(value[0])
        gene_interval_arrays(value, root_view)
        check_interval_active()

        url = items[2]
        params = dict(zip(items[0], items[1]))
        data_form = [url, params]
        # print(rf'added parames = {data_form}')
        Cons.finetree_parms_arrays.append(data_form)


# 2024.07.04: Operating normal mode
# (2024.07.25): NYX added to normal mode
def handle_normal_mode(event, tags, iden, title, root_view, tv, host, port):
    Cons.data_sending = True
    if 'checked' in tags:
        tv.item(iden, tags='unchecked')
    else:
        tv.item(iden, tags='checked')
    if Cons.selected_model in 'Uncooled':
        hex_value = select_item(event, root_view)
        ic('handle_normal_mode in MF', hex_value)
        Comm.send_cmd_for_uncooled(hex_value, title, root_view)
    elif Cons.selected_model == 'DRS':
        hex_array = select_item(event, root_view)
        Comm.send_cmd_for_drs(host, port, hex_array, root_view)
    elif Cons.selected_model == 'MiniGimbal':
        hex_array = select_item(event, root_view)
        print(rf'{datetime.now()} : {title}')
        # print(Cons.only_socket)
        # print(hex_array)
        Comm.send_to_mini(Cons.only_socket, hex_array)
    elif Cons.selected_model == 'NYX Series':
        send_data_for_nyx(event, root_view)
    elif Cons.selected_model == 'FineTree':
        index = int(iden.split('번')[0])
        items = Cons.fine_tree_cmd_data[index]
        print(items)
        url = items[2]
        params = dict(zip(items[0], items[1]))
        # for i, item in enumerate(items[0]):
        #     form = {rf'{item[0][i]}': rf'{item[1][i]}'}
        #     params.update(form)
        print(params)
        Comm.fine_tree_send_cgi(url, params)
    elif Cons.selected_model == 'Multi':
        hex_value = select_item(event, root_view)
        # print(hex_value)
        Comm.send_cmd_only_for_multi(hex_value)


def gene_interval_arrays(value, root_view):
    if not Cons.script_cmd_itv_arrs:
        Cons.script_cmd_itv_arrs = Cons.script_cmd_itv_arrs or [[0] * 2 for _ in range(len(Cons.script_cmd_titles))]
        for i, cmd_title in enumerate(Cons.script_cmd_titles):
            Cons.script_cmd_itv_arrs[i][0] = cmd_title
            print(Cons.script_cmd_itv_arrs[i])
    else:
        added = [value[0], 0]
        Cons.script_cmd_itv_arrs.append(added)
        print(f'added = {added}')
        print(Cons.script_cmd_itv_arrs)
        print(Cons.script_cmd_arrs)
    script_tb = tb.Table(root_view)


# (2024.07.04): Clear the Script Arrays
def clr_table_arrays(root):
    Cons.script_cmd_arrs = []
    Cons.script_cmd_titles = []
    Cons.script_itv_arrs = []
    Cons.script_cmd_itv_arrs = []
    tb.Table(root)


# PopUp when network textfield is empty
def show_network_dialog(root_view, dialog_text):
    # network_notification()
    dialog_txt = 'Please enter a network info.'
    dialog = Dialog.DialogBox(root_view, dialog_txt)


# Get the Command from csv file
# (2024.07.29) change protocol list base on a selected model
def get_data_from_csv(file_path) -> [(str, str)]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sel_model = Cons.selected_model
    command_data = []

    if sel_model in Cons.selected_model:
        # print(rf'get csv {sel_model}')
        sh = wb[f'{sel_model}']
    else:
        return command_data

    max_column = sh.max_column
    max_row = sh.max_row
    if sel_model in ['Uncooled', 'DRS', 'NYX Series', 'MiniGimbal', 'Multi', 'CTEC']:
        for i, row in enumerate(sh.iter_rows(max_col=max_column - 2, max_row=max_row - 2), start=3):
            cmd_title = sh.cell(row=i, column=2).value
            cmd_data = sh.cell(row=i, column=3).value
            cmd_pair = (cmd_title, cmd_data)
            command_data.append(cmd_pair)
    elif sel_model == 'FineTree':
        for i, row in enumerate(sh.iter_rows(max_col=max_column - 2, max_row=max_row - 2), start=3):
            title = sh.cell(row=i, column=2).value
            data = sh.cell(row=i, column=3).value
            para = sh.cell(row=i, column=4).value
            para_arr = str(para).split(',')
            value = sh.cell(row=i, column=5).value
            value_arr = str(value).split(',')
            url = sh.cell(row=i, column=6).value
            cmd_pair = (title, data)
            command_data.append(cmd_pair)

            fine_tree_cmd_data = (para_arr, value_arr, url)
            Cons.fine_tree_cmd_data.append(fine_tree_cmd_data)

    return command_data


# 2025.06.17: Added check whether the wininfo is exist
# (2024.07.05) Capture an Image from RTSP
# (2024.10.24): change a library to mss from pillow because pillow only detects the main monitor
def capture_image(root, filename):
    ti.sleep(1)
    try:
        if not root.winfo_exists():
            print('root not exists')
            return

        x = root.winfo_rootx()
        y = root.winfo_rooty()
        w = root.winfo_width()
        h = root.winfo_height()
        # print(rf'capture position is {x}, {y}, {w}, {h}')

        if w <= 0 or h <= 0:
            print('width or height is 0')
            return

        monitor = {
            "top": y,
            "left": x,
            "width": w,
            "height": h
        }
        with mss.mss() as sct:
            screenshot = sct.grab(monitor)
            mss.tools.to_png(screenshot.rgb, screenshot.size, output=filename)
    except Exception as e:
        print(f'capture image error: {e}')
# def do_capture(monitor_geom, filename, result_callback=None):
#     try:
#         top, left, width, height = monitor_geom
#         if width <= 0 or height <= 0:
#             raise ValueError("invalid size for capture")
#
#         monitor = {
#             "top": top,
#             "left": left,
#             "width": width,
#             "height": height
#         }
#         with mss.mss() as sct:
#             screenshot = sct.grab(monitor)
#             mss.tools.to_png(screenshot.rgb, screenshot.size, output=filename)
#         if result_callback:
#             result_callback(True, filename)
#     except Exception as e:
#         if result_callback:
#             result_callback(False, e)
#         else:
#             print(f'capture image error: {e}')
#
#
# # 메인 스레드에서 호출할 래퍼: tkinter API를 여기서만 사용
# def capture_image(root, filename, result_callback=None):
#     if not root.winfo_exists():
#         print('root not exists')
#         if result_callback:
#             result_callback(False, RuntimeError("root not exists"))
#         return
#
#     x = root.winfo_rootx()
#     y = root.winfo_rooty()
#     w = root.winfo_width()
#     h = root.winfo_height()
#
#     if w <= 0 or h <= 0:
#         print('width or height is 0')
#         if result_callback:
#             result_callback(False, RuntimeError("invalid geometry"))
#         return
#
#     # mss expects top/left/width/height in its own coordinate style
#     monitor_geom = (y, x, w, h)  # 기존 코드가 top=y, left=x
#     # 백그라운드에서 실제 캡처
#     thread = threading.Thread(target=do_capture, args=(monitor_geom, filename, result_callback), daemon=True)
#     thread.start()


def print_monitor_info():
    monitors = get_monitors()
    for i, monitor in enumerate(monitors):
        print(f"Monitor {i}:")
        print(f"  - X: {monitor.x}")
        print(f"  - Y: {monitor.y}")
        print(f"  - Width: {monitor.width}")
        print(f"  - Height: {monitor.height}")
        print()


# (2024.07.05) Get a Monitor Information
def get_secondary_monitor_bbox():
    monitors = get_monitors()
    if len(monitors) > 1:
        secondary_monitor = monitors[1]
        print(secondary_monitor.x, secondary_monitor.y,
              secondary_monitor.x + secondary_monitor.width,
              secondary_monitor.y + secondary_monitor.height)
        return (secondary_monitor.x, secondary_monitor.y,
                secondary_monitor.x + secondary_monitor.width,
                secondary_monitor.y + secondary_monitor.height)
    return None
